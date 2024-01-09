import time
import threading
import os
import openpyxl

from PyQt5 import QtCore, QtGui, QtWidgets
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from openpyxl.styles import Font


class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(400, 330)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(Form.sizePolicy().hasHeightForWidth())
        Form.setSizePolicy(sizePolicy)
        Form.setMinimumSize(QtCore.QSize(400, 330))
        Form.setMaximumSize(QtCore.QSize(400, 330))
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(10, 140, 81, 21))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(Form)
        self.pushButton_2.setGeometry(QtCore.QRect(10, 170, 81, 21))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(Form)
        self.pushButton_3.setGeometry(QtCore.QRect(10, 200, 81, 21))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(Form)
        self.pushButton_4.setGeometry(QtCore.QRect(10, 230, 81, 21))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_5 = QtWidgets.QPushButton(Form)
        self.pushButton_5.setGeometry(QtCore.QRect(10, 80, 81, 21))
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_6 = QtWidgets.QPushButton(Form)
        self.pushButton_6.setGeometry(QtCore.QRect(10, 110, 81, 21))
        self.pushButton_6.setObjectName("pushButton_6")
        self.pushButton_7 = QtWidgets.QPushButton(Form)
        self.pushButton_7.setGeometry(QtCore.QRect(10, 50, 81, 21))
        self.pushButton_7.setObjectName("pushButton_7")
        self.lineEdit = QtWidgets.QLineEdit(Form)
        self.lineEdit.setGeometry(QtCore.QRect(110, 270, 281, 20))
        self.lineEdit.setObjectName("lineEdit")
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(110, 51, 281, 20))
        self.label.setObjectName("label")
        self.line = QtWidgets.QFrame(Form)
        self.line.setGeometry(QtCore.QRect(90, 40, 21, 221))
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.line_2 = QtWidgets.QFrame(Form)
        self.line_2.setGeometry(QtCore.QRect(0, 250, 411, 20))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.pushButton_8 = QtWidgets.QPushButton(Form)
        self.pushButton_8.setGeometry(QtCore.QRect(10, 270, 81, 21))
        self.pushButton_8.setObjectName("pushButton_8")
        self.label_2 = QtWidgets.QLabel(Form)
        self.label_2.setGeometry(QtCore.QRect(110, 80, 281, 20))
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(Form)
        self.label_3.setGeometry(QtCore.QRect(110, 110, 281, 20))
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(Form)
        self.label_4.setGeometry(QtCore.QRect(110, 140, 281, 20))
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(Form)
        self.label_5.setGeometry(QtCore.QRect(110, 170, 281, 20))
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(Form)
        self.label_6.setGeometry(QtCore.QRect(110, 200, 281, 20))
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(Form)
        self.label_7.setGeometry(QtCore.QRect(110, 230, 281, 20))
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(Form)
        self.label_8.setGeometry(QtCore.QRect(110, 300, 281, 20))
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        self.line_3 = QtWidgets.QFrame(Form)
        self.line_3.setGeometry(QtCore.QRect(0, 30, 411, 20))
        self.line_3.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.lineEdit_2 = QtWidgets.QLineEdit(Form)
        self.lineEdit_2.setGeometry(QtCore.QRect(10, 10, 181, 20))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_3 = QtWidgets.QLineEdit(Form)
        self.lineEdit_3.setGeometry(QtCore.QRect(202, 10, 191, 20))
        self.lineEdit_3.setObjectName("lineEdit_3")

        self.retranslateUi(Form)
        self.pushButton_7.clicked.connect(self.call)
        self.pushButton_5.clicked.connect(self.call_1)
        self.pushButton_6.clicked.connect(self.call_2)
        self.pushButton.clicked.connect(self.call_3)
        self.pushButton_2.clicked.connect(self.call_4)
        self.pushButton_3.clicked.connect(self.call_5)
        self.pushButton_4.clicked.connect(self.call_6)
        self.pushButton_8.clicked.connect(self.call_7)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "B.Parser"))
        self.pushButton.setText(_translate("Form", "ОГМ (раб)"))
        self.pushButton_2.setText(_translate("Form", "МП (все)"))
        self.pushButton_3.setText(_translate("Form", "МП (раб)"))
        self.pushButton_4.setText(_translate("Form", "СИП (все)"))
        self.pushButton_5.setText(_translate("Form", "БпоНТ (норм)"))
        self.pushButton_6.setText(_translate("Form", "ОГМ (все)"))
        self.pushButton_7.setText(_translate("Form", "БпоНТ (все)"))
        self.label.setText(_translate("Form", ""))
        self.pushButton_8.setText(_translate("Form", "по ссылке"))
        self.label_2.setText(_translate("Form", ""))
        self.label_3.setText(_translate("Form", ""))
        self.label_4.setText(_translate("Form", ""))
        self.label_5.setText(_translate("Form", ""))
        self.label_6.setText(_translate("Form", ""))
        self.label_7.setText(_translate("Form", ""))
        self.label_8.setText(_translate("Form", ""))
        

    def call(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)


        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)           
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')                    
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)  
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')            
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_3.get(SRC)           
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')                    
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = 'http://portal/Pages/Phonebook.aspx#InplviewHashefe8c997-a328-44e4-9f3e-56b969c56b02=FilterField1%3DDivision-FilterValue1%3D%25D0%25A3%25D0%259D%25D0%259E%25D0%25A2%25D0%25B8%25D0%25A1%25D0%25A0' # нормировщики отз
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label.setText(f'in: {len(in_z)}, out: {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")')        
            name_i += 1    
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'БпоНТ_(все).xlsx'
        wb.save(save_path_1)


    def call_1(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)


        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'

                driver_3.get(SRC)
           
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')            
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = 'http://portal/Pages/Phonebook.aspx#InplviewHashefe8c997-a328-44e4-9f3e-56b969c56b02=FilterField1%3DDivision-FilterValue1%3D%25D0%25A3%25D0%259D%25D0%259E%25D0%25A2%25D0%25B8%25D0%25A1%25D0%25A0-FilterFields2%3DJobTitle-FilterValues2%3D%25D0%2592%25D0%25B5%25D0%25B4%25D1%2583%25D1%2589%25D0%25B8%25D0%25B9%2520%25D0%25B8%25D0%25BD%25D0%25B6%25D0%25B5%25D0%25BD%25D0%25B5%25D1%2580%2520%25D0%25BF%25D0%25BE%2520%25D0%25BD%25D0%25BE%25D1%2580%25D0%25BC%25D0%25B8%25D1%2580%25D0%25BE%25D0%25B2%25D0%25B0%25D0%25BD%25D0%25B8%25D1%258E%2520%25D1%2582%25D1%2580%25D1%2583%25D0%25B4%25D0%25B0%253B%2523%25D0%2598%25D0%25BD%25D0%25B6%25D0%25B5%25D0%25BD%25D0%25B5%25D1%2580%2520%25D0%25BF%25D0%25BE%2520%25D0%25BD%25D0%25BE%25D1%2580%25D0%25BC%25D0%25B8%25D1%2580%25D0%25BE%25D0%25B2%25D0%25B0%25D0%25BD%25D0%25B8%25D1%258E%2520%25D1%2582%25D1%2580%25D1%2583%25D0%25B4%25D0%25B0'
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label_2.setText(f'in: {len(in_z)}, out: {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")')
            name_i += 1
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'БпоНТ_(норм).xlsx'
        wb.save(save_path_1)


    def call_2(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)


        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_3.get(SRC)
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}') 
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = 'http://portal/Pages/Phonebook.aspx#InplviewHashefe8c997-a328-44e4-9f3e-56b969c56b02=p_Division%3D%25d0%259a%25d0%259f-FolderCTID%3D0x012001-FilterField1%3DDivision-FilterValue1%3D%25D0%259E%25D0%2593%25D0%259C' # огм
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label_3.setText(f'in: {len(in_z)}, out {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")')
            name_i += 1
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'ОГМ_(все).xlsx'
        wb.save(save_path_1)


    def call_3(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)


        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_3.get(SRC)
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = 'http://portal/Pages/Phonebook.aspx#InplviewHashefe8c997-a328-44e4-9f3e-56b969c56b02=SortField%3DDivision-SortDir%3DDesc-FilterField1%3DDivision-FilterValue1%3D%25D0%259E%25D0%2593%25D0%259C-FilterFields2%3DJobTitle-FilterValues2%3D%25D0%25AD%25D0%25BB%25D0%25B5%25D0%25BA%25D1%2582%25D1%2580%25D0%25BE%25D0%25BC%25D0%25BE%25D0%25BD%25D1%2582%25D0%25B5%25D1%2580%2520%25D0%25BF%25D0%25BE%2520%25D1%2580%25D0%25B5%25D0%25BC%25D0%25BE%25D0%25BD%25D1%2582%25D1%2583%2520%25D0%25B8%2520%25D0%25BE%25D0%25B1%25D1%2581%25D0%25BB%25D1%2583%25D0%25B6%25D0%25B8%25D0%25B2%25D0%25B0%25D0%25BD%25D0%25B8%25D1%258E%2520%25D1%258D%25D0%25BB%25D0%25B5%25D0%25BA%25D1%2582%25D1%2580%25D0%25BE%25D0%25BE%25D0%25B1%25D0%25BE%25D1%2580%25D1%2583%25D0%25B4%25D0%25BE%25D0%25B2%25D0%25B0%25D0%25BD%25D0%25B8%25D1%258F%253B%2523%25D0%25AD%25D0%25BB%25D0%25B5%25D0%25BA%25D1%2582%25D1%2580%25D0%25BE%25D0%25BC%25D0%25B5%25D1%2585%25D0%25B0%25D0%25BD%25D0%25B8%25D0%25BA%2520%25D0%25BF%25D0%25BE%2520%25D0%25BB%25D0%25B8%25D1%2584%25D1%2582%25D0%25B0%25D0%25BC%253B%2523%25D0%25A8%25D0%25BB%25D0%25B8%25D1%2584%25D0%25BE%25D0%25B2%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25A4%25D1%2580%25D0%25B5%25D0%25B7%25D0%25B5%25D1%2580%25D0%25BE%25D0%25B2%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25A2%25D0%25BE%25D0%25BA%25D0%25B0%25D1%2580%25D1%258C%252D%25D1%2580%25D0%25B0%25D1%2581%25D1%2582%25D0%25BE%25D1%2587%25D0%25BD%25D0%25B8%25D0%25BA%253B%2523%25D0%25A2%25D0%25BE%25D0%25BA%25D0%25B0%25D1%2580%25D1%258C%253B%2523%25D0%25A1%25D0%25BB%25D0%25B5%25D1%2581%25D0%25B0%25D1%2580%25D1%258C%252D%25D1%2580%25D0%25B5%25D0%25BC%25D0%25BE%25D0%25BD%25D1%2582%25D0%25BD%25D0%25B8%25D0%25BA%253B%2523%25D0%25A0%25D0%25B5%25D0%25B7%25D1%2587%25D0%25B8%25D0%25BA%2520%25D0%25BD%25D0%25B0%2520%25D0%25BF%25D0%25B8%25D0%25BB%25D0%25B0%25D1%2585%252C%2520%25D0%25BD%25D0%25BE%25D0%25B6%25D0%25BE%25D0%25B2%25D0%25BA%25D0%25B0%25D1%2585%2520%25D0%25B8%2520%25D1%2581%25D1%2582%25D0%25B0%25D0%25BD%25D0%25BA%25D0%25B0%25D1%2585%253B%2523%25D0%259F%25D1%2580%25D1%2583%25D0%25B6%25D0%25B8%25D0%25BD%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%259E%25D0%25BF%25D0%25B5%25D1%2580%25D0%25B0%25D1%2582%25D0%25BE%25D1%2580%2520%25D1%2581%25D1%2582%25D0%25B0%25D0%25BD%25D0%25BA%25D0%25BE%25D0%25B2%2520%25D1%2581%2520%25D0%25BF%25D1%2580%25D0%25BE%25D0%25B3%25D1%2580%25D0%25B0%25D0%25BC%25D0%25BC%25D0%25BD%25D1%258B%25D0%25BC%2520%25D1%2583%25D0%25BF%25D1%2580%25D0%25B0%25D0%25B2%25D0%25BB%25D0%25B5%25D0%25BD%25D0%25B8%25D0%25B5%25D0%25BC%253B%2523%25D0%259D%25D0%25B0%25D0%25BB%25D0%25B0%25D0%25B4%25D1%2587%25D0%25B8%25D0%25BA%2520%25D0%25BC%25D0%25B0%25D1%2588%25D0%25B8%25D0%25BD%2520%25D0%25B8%2520%25D0%25B0%25D0%25B2%25D1%2582%25D0%25BE%25D0%25BC%25D0%25B0%25D1%2582%25D0%25B8%25D1%2587%25D0%25B5%25D1%2581%25D0%25BA%25D0%25B8%25D1%2585%2520%25D0%25BB%25D0%25B8%25D0%25BD%25D0%25B8%25D0%25B9%2520%25D0%25BF%25D0%25BE%2520%25D0%25BF%25D1%2580%25D0%25BE%25D0%25B8%25D0%25B7%25D0%25B2%25D0%25BE%25D0%25B4%25D1%2581%25D1%2582%25D0%25B2%25D1%2583%2520%25D0%25B8%25D0%25B7%25D0%25B4%25D0%25B5%25D0%25BB%25D0%25B8%25D0%25B9%2520%25D0%25B8%25D0%25B7%2520%25D0%25BF%25D0%25BB%25D0%25B0%25D1%2581%25D1%2582%25D0%25BC%25D0%25B0%25D1%2581%25D1%2581'
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label_4.setText(f'in: {len(in_z)}, out: {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75        
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")')
            name_i += 1
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'ОГМ_(раб).xlsx'
        wb.save(save_path_1)


    def call_4(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)


        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_3.get(SRC)
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = 'http://portal/Pages/Phonebook.aspx#InplviewHashefe8c997-a328-44e4-9f3e-56b969c56b02=SortField%3DDivision-SortDir%3DDesc-FilterField1%3DDivision-FilterValue1%3D%25D0%259C%25D0%259F' # мп (все)
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label_5.setText(f'in: {len(in_z)}, out: {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")')
            name_i += 1
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'МП_(все).xlsx'
        wb.save(save_path_1)


    def call_5(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)


        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'

                driver_3.get(SRC)
           
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')               
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = 'http://portal/Pages/Phonebook.aspx#InplviewHashefe8c997-a328-44e4-9f3e-56b969c56b02=SortField%3DDivision-SortDir%3DDesc-FilterField1%3DDivision-FilterValue1%3D%25D0%259C%25D0%259F-FilterFields2%3DJobTitle-FilterValues2%3D%25D0%25AD%25D0%25BB%25D0%25B5%25D0%25BA%25D1%2582%25D1%2580%25D0%25BE%25D0%25B3%25D0%25B0%25D0%25B7%25D0%25BE%25D1%2581%25D0%25B2%25D0%25B0%25D1%2580%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25AD%25D0%25BB%25D0%25B5%25D0%25BA%25D1%2582%25D1%2580%25D0%25BE%25D0%25BC%25D0%25BE%25D0%25BD%25D1%2582%25D0%25B5%25D1%2580%2520%25D0%25BF%25D0%25BE%2520%25D1%2580%25D0%25B5%25D0%25BC%25D0%25BE%25D0%25BD%25D1%2582%25D1%2583%2520%25D0%25B8%2520%25D0%25BE%25D0%25B1%25D1%2581%25D0%25BB%25D1%2583%25D0%25B6%25D0%25B8%25D0%25B2%25D0%25B0%25D0%25BD%25D0%25B8%25D1%258E%2520%25D1%258D%25D0%25BB%25D0%25B5%25D0%25BA%25D1%2582%25D1%2580%25D0%25BE%25D0%25BE%25D0%25B1%25D0%25BE%25D1%2580%25D1%2583%25D0%25B4%25D0%25BE%25D0%25B2%25D0%25B0%25D0%25BD%25D0%25B8%25D1%258F%253B%2523%25D0%25A8%25D0%25BB%25D0%25B8%25D1%2584%25D0%25BE%25D0%25B2%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25A4%25D1%2580%25D0%25B5%25D0%25B7%25D0%25B5%25D1%2580%25D0%25BE%25D0%25B2%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25A2%25D0%25BE%25D0%25BA%25D0%25B0%25D1%2580%25D1%258C%252D%25D1%2580%25D0%25B5%25D0%25B2%25D0%25BE%25D0%25BB%25D1%258C%25D0%25B2%25D0%25B5%25D1%2580%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25A2%25D0%25BE%25D0%25BA%25D0%25B0%25D1%2580%25D1%258C%252D%25D1%2580%25D0%25B0%25D1%2581%25D1%2582%25D0%25BE%25D1%2587%25D0%25BD%25D0%25B8%25D0%25BA%253B%2523%25D0%25A2%25D0%25BE%25D0%25BA%25D0%25B0%25D1%2580%25D1%258C%252D%25D0%25BA%25D0%25B0%25D1%2580%25D1%2583%25D1%2581%25D0%25B5%25D0%25BB%25D1%258C%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25A2%25D0%25BE%25D0%25BA%25D0%25B0%25D1%2580%25D1%258C%253B%2523%25D0%25A2%25D0%25B5%25D1%2580%25D0%25BC%25D0%25B8%25D1%2581%25D1%2582%253B%2523%25D0%25A1%25D1%2582%25D1%2580%25D0%25BE%25D0%25B3%25D0%25B0%25D0%25BB%25D1%258C%25D1%2589%25D0%25B8%25D0%25BA%253B%2523%25D0%25A1%25D0%25BB%25D0%25B5%25D1%2581%25D0%25B0%25D1%2580%25D1%258C%252D%25D1%2581%25D0%25B0%25D0%25BD%25D1%2582%25D0%25B5%25D1%2585%25D0%25BD%25D0%25B8%25D0%25BA%253B%2523%25D0%25A1%25D0%25BB%25D0%25B5%25D1%2581%25D0%25B0%25D1%2580%25D1%258C%252D%25D1%2580%25D0%25B5%25D0%25BC%25D0%25BE%25D0%25BD%25D1%2582%25D0%25BD%25D0%25B8%25D0%25BA%253B%2523%25D0%25A1%25D0%25BB%25D0%25B5%25D1%2581%25D0%25B0%25D1%2580%25D1%258C%252D%25D0%25B8%25D0%25BD%25D1%2581%25D1%2582%25D1%2580%25D1%2583%25D0%25BC%25D0%25B5%25D0%25BD%25D1%2582%25D0%25B0%25D0%25BB%25D1%258C%25D1%2589%25D0%25B8%25D0%25BA'
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label_6.setText(f'in: {len(in_z)}, out: {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")')    
            name_i += 1      
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'МП_(раб).xlsx'
        wb.save(save_path_1)


    def call_6(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)
            

        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}') 
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')     
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_3.get(SRC)
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')           
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = 'http://portal/Pages/Phonebook.aspx#InplviewHashefe8c997-a328-44e4-9f3e-56b969c56b02=FilterField1%3DDivision-FilterValue1%3D%25D0%25A6%25D0%2595%25D0%25A5%252023' # сип
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label_7.setText(f'in: {len(in_z)}, out: {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")') 
            name_i += 1
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'23_цех.xlsx'
        wb.save(save_path_1)


    def call_7(self):
        def resource_path(r_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.dirname(__file__)
            return os.path.join(base_path, r_path)

        
        def insert_reg(url):
            chrome_log = f'{log}:{passs}@'
            url_segment_1 = url[:7]
            url_segment_2 = url[7:]
            SRC = f'{url_segment_1}{chrome_log}{url_segment_2}'
            return SRC


        def main(page_url):
            global driver
            service = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(page_url)
            time.sleep(2)

        def links_get():
            global driver
            iframe = driver.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/div[2]/div/div[4]/div/div[3]')#.get_attribute('href') #'/html/body/form/div[5]')
            hrefs = iframe.find_elements(By.TAG_NAME, 'a')
            for i in hrefs:
                try:
                    if i.get_attribute('href').split('15')[0] == 'http://portal/_layouts/':
                        links.append(i.get_attribute('href'))
                    else:
                        continue
                except:
                    continue


        def try_next_page():
            global driver
            button_1 = driver.find_element(By.ID, 'pagingWPQ2next')
            button_11 = button_1.find_element(By.TAG_NAME, 'a')
            button_11.click()


        def scrab_profile_1(links_segment):
            service_1 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_1 = webdriver.Chrome(service=service_1, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_1.get(SRC)
                time.sleep(1)
                iframe_1 = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_1.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_1.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_1.get(iframe_2)
                    iframe_3 = driver_1.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_1.quit()


        def scrab_profile_2(links_segment):
            service_2 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_2 = webdriver.Chrome(service=service_2, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_2.get(SRC)
                time.sleep(1)
                iframe_1 = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_2.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_2.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_2.get(iframe_2)
                    iframe_3 = driver_2.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')
                n += 1
            driver_2.quit()


        def scrab_profile_3(links_segment):
            service_3 = Service(executable_path=resource_path('./driver/chromedriver.exe'))
            driver_3 = webdriver.Chrome(service=service_3, options=options)
            n = 1
            for j in links_segment:
                chrome_log = f'{log}:{passs}@'
                src_segment_1 = j[:7]
                src_segment_2 = j[7:]
                SRC = f'{src_segment_1}{chrome_log}{src_segment_2}'
                driver_3.get(SRC)
                time.sleep(1)
                iframe_1 = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[1]/td/div')
                text_name = iframe_1.text
                iframe_prof = driver_3.find_element(By.XPATH, '/html/body/form/div[5]/div/div/div/span/div[1]/table/tbody/tr/td/div/div/div/div/div[1]/table/tbody/tr[1]/td[2]/table/tbody/tr[5]/td[2]/div')
                text_prof = iframe_prof.text
                try:
                    iframe_2 = driver_3.find_element(By.TAG_NAME, 'iframe').get_attribute('src')
                    driver_3.get(iframe_2)
                    iframe_3 = driver_3.find_element(By.XPATH, '/html/body/table/tbody/tr/td').get_attribute('bgcolor') #'/html/body/form/div[5]')
                    if iframe_3 == 'MediumSeaGreen':
                        iframe_3 = 'НА ТЕРРИТОРИИ'
                        in_z.append('1')
                    elif iframe_3 == 'Tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    elif iframe_3 == 'tomato':
                        iframe_3 = 'ОТСУТСТВУЕТ'
                        out_z.append('1')
                    full_list.append(f'{text_name}:{iframe_3}:{time.strftime("%X")}:{text_prof}')
                except:
                    no_z.append('1')
                    full_list.append(f'{text_name}:НЕ РАБОТАЕТ:{time.strftime("%X")}:{text_prof}')          
                n += 1
            driver_3.quit()
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--blink-settings=imagesEnabled=false')
        log = self.lineEdit_2.text()
        passs = self.lineEdit_3.text()
        input_url = self.lineEdit.text()
        insert_input_url = insert_reg(input_url)
        main(insert_input_url)
        time.sleep(3)
        links = []
        in_z = []
        out_z = []
        no_z = []
        full_list = []
        time_list = []
        x = 0
        tumbler_1 = True
        while (tumbler_1):
            links_get()
            x += 1
            try:
                try_next_page()
                time.sleep(1)
            except:
                tumbler_1 = False
                continue
        driver.quit()
        num_segment = (len(links) / 3) + 1
        int_num_segment = int(num_segment)
        chunk_size = int_num_segment
        chunks = [links[k:k + chunk_size] for k in range(0, len(links), chunk_size)]
        t1 = threading.Thread(target=scrab_profile_1, args=(chunks[0],), daemon=True)
        t2 = threading.Thread(target=scrab_profile_2, args=(chunks[1],), daemon=True)
        t3 = threading.Thread(target=scrab_profile_3, args=(chunks[2],), daemon=True)
        time_start = time.strftime('%X')
        t1.start()
        t2.start()
        t3.start()
        t1.join()    
        t2.join()
        t3.join()
        self.label_8.setText(f'in: {len(in_z)}, out: {len(out_z)+len(no_z)}. Файл готов. {time_start}-{time.strftime("%X")}')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 75
        name_i = int(1)
        for name in full_list:
            name_1 = name.split(':')
            exec(f'ws["A{name_i}"] = name_1[0]')
            exec(f'ws["B{name_i}"] = name_1[1]')
            exec(f'ws["C{name_i}"] = name_1[5]')
            exec(f'ws["D{name_i}"] = "{name_1[2]}:{name_1[3]}"')
            if name_1[1] == 'НА ТЕРРИТОРИИ':
                exec(f'ws["B{name_i}"].font = Font(color = "0F8500")')
            if name_1[1] == 'ОТСУТСТВУЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF0000")')
            if name_1[1] == 'НЕ РАБОТАЕТ':
                exec(f'ws["B{name_i}"].font = Font(color = "FF8400")')
            name_i += 1
        ws.name = 'Список'
        save_path = os.path.dirname(__file__)
        save_path_1 = f'Своя_ссылка.xlsx'
        wb.save(save_path_1)


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())
